from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl
import pathlib

root = Tk()
root.title("✈FADE")
root.geometry('700x400+300+200')
root.resizable(False, False)
root.configure(bg="#326273")

# Check if the Excel file exists, create it if it doesn't
file = pathlib.Path('Backened_data.xlsx')
if not file.exists():
    file = openpyxl.Workbook()
    sheet = file.active
    sheet['A1'] = "Full name"
    sheet['B1'] = "Age"
    sheet['C1'] = "IT No"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"
    file.save('Backened_data.xlsx')


def submit():
    name = namevalue.get()
    age = agevalue.get()
    ITno = ITNovalue.get()
    gender = gender_combobox.get()
    address = addressEntry.get(1.0, END)

    # Check for empty fields
    if not name or not age or not ITno or not gender or not address.strip():
        messagebox.showwarning("Input Error", "Please fill out all fields.")
        return

    try:
        # Load existing Excel file
        file = openpyxl.load_workbook('Backened_data.xlsx')
        sheet = file.active

        # Insert the form data into the sheet
        next_row = sheet.max_row + 1
        sheet[f'A{next_row}'] = name
        sheet[f'B{next_row}'] = age
        sheet[f'C{next_row}'] = ITno
        sheet[f'D{next_row}'] = gender
        sheet[f'E{next_row}'] = address.strip()  # Remove any extra spaces

        # Save the file
        file.save('Backened_data.xlsx')

        messagebox.showinfo("Submission", "Form data has been saved successfully.")
        # Clear the form fields
        namevalue.set('')
        agevalue.set('')
        ITNovalue.set('')
        addressEntry.delete(1.0, END)
    
    except PermissionError:
        messagebox.showerror("Permission Error", "Permission denied while trying to save the file. Please ensure the file is not open elsewhere or check your file permissions.")


def clear():
    namevalue.set('')
    agevalue.set('')
    ITNovalue.set('')
    addressEntry.delete(1.0, END)


# Heading
Label(root, text="PLEASE FILL OUT THIS ENTRY FORM:", font="arial 13", bg="#326273", fg="#fff").place(x=20, y=20)

# Labels
Label(root, text='Name', font=23, bg="#326273", fg="#fff").place(x=50, y=100)
Label(root, text='Age', font=23, bg="#326273", fg="#fff").place(x=50, y=150)
Label(root, text='IT No.', font=23, bg="#326273", fg="#fff").place(x=50, y=200)
Label(root, text='Gender', font=23, bg="#326273", fg="#fff").place(x=370, y=200)
Label(root, text='Address', font=23, bg="#326273", fg="#fff").place(x=50, y=250)

# Entry fields
namevalue = StringVar()
agevalue = StringVar()
ITNovalue = StringVar()

nameEntry = Entry(root, textvariable=namevalue, width=45, bd=2, font=20)
ageEntry = Entry(root, textvariable=agevalue, width=45, bd=2, font=20)
ITNoEntry = Entry(root, textvariable=ITNovalue, width=15, bd=2, font=20)

# Gender Combobox
gender_combobox = Combobox(root, values=['Male', 'Female', 'None'], font='arial 14', state='r', width=14)
gender_combobox.place(x=440, y=200)
gender_combobox.set('')

addressEntry = Text(root, width=50, height=4, bd=2)

# Place widgets
nameEntry.place(x=150, y=100)
ageEntry.place(x=150, y=150)
ITNoEntry.place(x=150, y=200)
addressEntry.place(x=200, y=250)

# Buttons
Button(root, text="Submit", bg="#fff", fg="#346273", width=15, height=2, command=submit).place(x=200, y=350)
Button(root, text="Clear", bg="#fff", fg="#346273", width=15, height=2, command=clear).place(x=340, y=350)
Button(root, text="Exit", bg="#fff", fg="#346273", width=15, height=2, command=lambda: root.destroy()).place(x=480, y=350)

root.mainloop()
